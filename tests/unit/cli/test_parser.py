"""Test the CLI parser."""

# ruff: noqa: ANN001, ANN201
from unittest.mock import patch
from zipfile import BadZipFile

from geneweaver.client.cli.alpha.parse import cli
from openpyxl.utils.exceptions import InvalidFileException
from typer.testing import CliRunner

runner = CliRunner()


@patch(
    "geneweaver.client.parser.general.get_headers",
    return_value=(["header1", "header2", "header3"], 0),
)
def test_get_headers(mock_get_headers):
    """Test the get_headers CLI command."""
    result = runner.invoke(cli, ["utils", "get-headers", "fake_path"])
    print(result.output)

    # # Check the output message
    assert "header1" in result.output
    assert "header2" in result.output
    assert "header3" in result.output
    # # Check the exit code
    assert result.exit_code == 0


@patch(
    "geneweaver.client.parser.general.get_headers",
    side_effect=ValueError("Unsupported file type: .txt"),
)
def test_get_headers_with_value_error(mock_get_headers):
    """Test the get_headers CLI command with a ValueError."""
    # Simulate the CLI execution
    result = runner.invoke(cli, ["utils", "get-headers", "dummy_path"])

    # Check the output message
    assert "Unsupported file type:" in result.output
    # Check the exit code
    assert result.exit_code == 1

    # Check if the mocked function was called
    mock_get_headers.assert_called_once()


@patch(
    "geneweaver.client.cli.alpha.parse.utils.xlsx.get_sheet_names",
    return_value=(["sheet1", "sheet2", "sheet3"]),
)
def test_get_sheet_names(mock_get_sheet_names):
    """Test the get_sheet_names CLI command."""
    # Simulate the CLI execution
    result = runner.invoke(cli, ["utils", "get-sheet-names", "fake_path"])

    # # Check the output message
    assert "sheet1" in result.output
    assert "sheet2" in result.output
    assert "sheet3" in result.output

    # Check the exit code
    assert result.exit_code == 0


@patch(
    "geneweaver.client.cli.alpha.parse.utils.xlsx.get_sheet_names",
    return_value=(["sheet1", "sheet2", "sheet3"]),
    side_effect=ValueError("Invalid File."),
)
def test_get_sheet_names_error(mock_get_sheet_names):
    """Test the get_sheet_names CLI command with an error."""
    # Simulate the CLI execution
    result = runner.invoke(cli, ["utils", "get-sheet-names", "fake_path"])

    # # Check the output message
    assert "Invalid File." in result.output

    # Check the exit code
    assert result.exit_code == 1


@patch(
    "geneweaver.client.cli.alpha.parse.utils.xlsx.load_workbook",
    return_value=(["sheet1", "sheet2", "sheet3"]),
    side_effect=InvalidFileException("Invalid File."),
)
def test_get_sheet_names_load_workbook_error(mock_get_sheet_names):
    """Test the get_sheet_names CLI command with an InvalidFileException error."""
    # Simulate the CLI execution
    result = runner.invoke(cli, ["utils", "get-sheet-names", "fake_path"])

    # # Check the output message
    assert "Invalid File." in result.output

    # Check the exit code
    assert result.exit_code == 1


@patch(
    "geneweaver.client.cli.alpha.parse.utils.xlsx.load_workbook",
    return_value=(["sheet1", "sheet2", "sheet3"]),
    side_effect=BadZipFile("Not a Zip File."),
)
def test_get_sheet_names_load_workbook_bad_zip_error(mock_get_sheet_names):
    """Test the get_sheet_names CLI command with a BadZipFile error."""
    # Simulate the CLI execution
    result = runner.invoke(cli, ["utils", "get-sheet-names", "fake_path"])

    # # Check the output message
    assert "Not a Zip File." in result.output

    # Check the exit code
    assert result.exit_code == 1
