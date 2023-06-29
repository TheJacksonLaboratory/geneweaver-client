"""Shared fixtures for the parser tests."""
# ruff: noqa: B905
from pathlib import Path
from typing import List, Tuple

import pytest
from openpyxl import Workbook, load_workbook

EXAMPLE_SHEET_NAMES = [
    ["Sheet1", "TestSheet", "AnotherSheet"],
    ["S1", "S2", "S3"],
    ["Sheet1", "Sheet2", "Sheet3", "Sheet4"],
    ["LPSF-Liver_topTable_no_outliers"],
]


@pytest.fixture(params=EXAMPLE_SHEET_NAMES)
def multi_sheet_excel_file(
    tmp_path: Path, request: pytest.FixtureRequest
) -> Tuple[Path, List[str]]:
    """Fixture for creating an Excel file with predefined sheet names."""
    # Create new workbook and add sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet created
    for sheet in request.param:
        wb.create_sheet(sheet)

    # Save the workbook to a temp file
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)

    return file_path, request.param


@pytest.fixture()
def not_an_xlsx_file(tmp_path: Path) -> Path:
    """Fixture for creating an Excel file with predefined sheet names."""
    file_path = tmp_path / "test.xlsx"
    file_path.touch()
    return file_path


EXAMPLE_DATA = [
    # Header exists at the first row
    ([["Header1", "Header2"], [1, 2], [3, 4]], 0, True),
    # Header exists at the second row
    ([["Not a Header"], ["Header1", "Header2"], [1, 2]], 1, True),
    # No header
    ([[1, 2], [3, 4], [5, 6]], -1, False),
    # Real Example Data
    (
        [
            ["Gene", "baseMean", "Log2FoldChange", "lfcSE", "stat", "p-value"],
            [
                "Amotl2",
                "241.5118269",
                "0.426986377",
                "0.079435004",
                "5.375292477",
                "7.65E-08",
            ],
        ],
        0,
        True,
    ),
    (
        [
            ["Differentially expressed"],
            ["Gene", "baseMean", "Log2FoldChange", "lfcSE", "stat", "p-value"],
            [
                "Amotl2",
                "241.5118269",
                "0.426986377",
                "0.079435004",
                "5.375292477",
                "7.65E-08",
            ],
        ],
        1,
        True,
    ),
    (
        [
            [],
            ["Differentially expressed"],
            ["Gene", "baseMean", "Log2FoldChange", "lfcSE", "stat", "p-value"],
            [
                "Amotl2",
                "241.5118269",
                "0.426986377",
                "0.079435004",
                "5.375292477",
                "7.65E-08",
            ],
        ],
        2,
        True,
    ),
]


@pytest.fixture(params=EXAMPLE_DATA)
def multi_sheet_excel_file_w_data(
    multi_sheet_excel_file: Tuple[Path, List[str]], request: pytest.FixtureRequest
) -> Tuple[Path, list, list]:
    """Fixture for creating an Excel file with predefined sheet names."""
    file_path, sheets = multi_sheet_excel_file
    workbook = load_workbook(file_path)
    for sheet_name in sheets:
        ws = workbook[sheet_name]
        for row in request.param[0]:
            ws.append(row)
    workbook.save(file_path)
    return file_path, sheets, request.param


@pytest.fixture()
def multi_sheet_excel_file_w_data_as_dict(
    multi_sheet_excel_file_w_data: Tuple[Path, list, list]
) -> Tuple[Path, list, int, list]:
    """Fixture for creating expected dict data from example data."""
    file_path, sheets, data = multi_sheet_excel_file_w_data
    data, header_index, has_header = data
    if not has_header:
        header_index = 0

    header = data[header_index]
    dicts = [dict(zip(header, data[i])) for i in range(header_index + 1, len(data))]

    return file_path, sheets, header_index, dicts
